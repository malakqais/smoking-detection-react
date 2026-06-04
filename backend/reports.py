"""Excel report generation for staff panel."""
import datetime
import io
from config import DB_PATH


def _connect():
    import sqlite3
    return sqlite3.connect(DB_PATH)


def _style_workbook(wb):
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    brand_fill = PatternFill("solid", fgColor="1E293B")
    brand_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="DC2626")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "brand_fill": brand_fill,
        "brand_font": brand_font,
        "header_fill": header_fill,
        "header_font": header_font,
        "border": border,
        "title_font": Font(name="Calibri", size=14, bold=True, color="1E293B"),
        "wrap": Alignment(wrap_text=True, vertical="top"),
    }


def build_user_report_xlsx(uid):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    conn = _connect()
    user = conn.execute(
        "SELECT id, name, email, role, created_at FROM users WHERE id = ?", (uid,)
    ).fetchone()
    if not user:
        conn.close()
        return None
    uid_, name, email, role, created_at = user
    violations = conn.execute(
        """SELECT id, timestamp, location, person_name, detected_type, image_path
           FROM violations WHERE person_name = ? OR user_id = ?
           ORDER BY timestamp DESC""",
        (name, uid_),
    ).fetchall()
    top_loc = conn.execute(
        """SELECT location, COUNT(*) FROM violations
           WHERE person_name = ? OR user_id = ?
           GROUP BY location ORDER BY COUNT(*) DESC LIMIT 1""",
        (name, uid_),
    ).fetchone()
    conn.close()

    wb = Workbook()
    styles = _style_workbook(wb)
    ws = wb.active
    ws.title = "User Report"

    ws.merge_cells("A1:F1")
    ws["A1"] = "SMOKEDET — USER VIOLATION REPORT"
    ws["A1"].fill = styles["brand_fill"]
    ws["A1"].font = styles["brand_font"]
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    meta = [
        ("Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Account name", name),
        ("Email", email),
        ("Role", role),
        ("Joined", created_at or "—"),
        ("Account ID", f"UID-{uid_:04d}"),
        ("Total violations", len(violations)),
        ("Top location", top_loc[0] if top_loc else "—"),
    ]
    row = 3
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    headers = ["#", "Violation ID", "Timestamp", "Location", "Detected type", "Evidence path"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["border"]
    header_row = row
    row += 1

    for i, v in enumerate(violations, 1):
        ws.cell(row=row, column=1, value=i).border = styles["border"]
        ws.cell(row=row, column=2, value=v[0]).border = styles["border"]
        ws.cell(row=row, column=3, value=v[1]).border = styles["border"]
        ws.cell(row=row, column=4, value=v[2]).border = styles["border"]
        ws.cell(row=row, column=5, value=v[3] or v[4]).border = styles["border"]
        ws.cell(row=row, column=6, value=v[5] or "").border = styles["border"]
        row += 1

    if not violations:
        ws.merge_cells(start_row=header_row + 1, start_column=1, end_row=header_row + 1, end_column=6)
        ws.cell(row=header_row + 1, column=1, value="No violations on record for this user.")

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["C"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = "".join(c if c.isalnum() or c in "._-" else "_" for c in name)
    filename = f"SmokeDet_User_Report_{safe_name}_{datetime.date.today().isoformat()}.xlsx"
    return buf.getvalue(), filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def build_platform_workbook():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font

    conn = _connect()
    users = conn.execute(
        """SELECT id, name, email, role, created_at, account_status,
                  (SELECT COUNT(*) FROM violations v WHERE v.person_name = users.name OR v.user_id = users.id)
           FROM users ORDER BY id"""
    ).fetchall()
    violations = conn.execute(
        """SELECT id, timestamp, location, person_name, detected_type, image_path
           FROM violations ORDER BY id DESC LIMIT 5000"""
    ).fetchall()
    disputes = conn.execute(
        """SELECT d.id, d.status, d.created_at, u.name, d.violation_id
           FROM violation_disputes d JOIN users u ON u.id = d.user_id
           ORDER BY d.created_at DESC LIMIT 500"""
    ).fetchall()
    conn.close()

    wb = Workbook()
    styles = _style_workbook(wb)

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "SMOKEDET PLATFORM REPORT"
    ws["A1"].font = styles["title_font"]
    ws["A3"] = "Generated"
    ws["B3"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Total users"
    ws["B4"] = len(users)
    ws["A5"] = "Total violations"
    ws["B5"] = len(violations)
    ws["A6"] = "Open disputes"
    ws["B6"] = sum(1 for d in disputes if d[1] in ("pending_admin", "awaiting_supervisor"))

    # Users sheet
    wu = wb.create_sheet("Users")
    uh = ["ID", "Name", "Email", "Role", "Status", "Joined", "Violations"]
    for c, h in enumerate(uh, 1):
        cell = wu.cell(row=1, column=c, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
    for r, u in enumerate(users, 2):
        wu.cell(row=r, column=1, value=u[0])
        wu.cell(row=r, column=2, value=u[1])
        wu.cell(row=r, column=3, value=u[2])
        wu.cell(row=r, column=4, value=u[3])
        wu.cell(row=r, column=5, value=u[5] or "active")
        wu.cell(row=r, column=6, value=u[4])
        wu.cell(row=r, column=7, value=u[6])

    # Violations sheet
    wv = wb.create_sheet("Violations")
    vh = ["ID", "Timestamp", "Location", "Person", "Type", "Image"]
    for c, h in enumerate(vh, 1):
        cell = wv.cell(row=1, column=c, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
    for r, v in enumerate(violations, 2):
        for c, val in enumerate(v, 1):
            wv.cell(row=r, column=c, value=val)

    wd = wb.create_sheet("Disputes")
    dh = ["ID", "Violation", "User", "Status", "Filed at"]
    for c, h in enumerate(dh, 1):
        cell = wd.cell(row=1, column=c, value=h)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
    for r, d in enumerate(disputes, 2):
        wd.cell(row=r, column=1, value=d[0])
        wd.cell(row=r, column=2, value=d[4])
        wd.cell(row=r, column=3, value=d[3])
        wd.cell(row=r, column=4, value=d[1])
        wd.cell(row=r, column=5, value=d[2])

    for sheet in (wu, wv, wd):
        for col in range(1, 10):
            sheet.column_dimensions[get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"SmokeDet_Platform_Report_{datetime.date.today().isoformat()}.xlsx"
    return buf.getvalue(), filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
